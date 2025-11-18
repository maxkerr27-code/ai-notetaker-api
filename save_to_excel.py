# save_to_excel.py
# Appends extracted JSON data to Excel, fixes headers, and auto-formats columns

import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# File names
json_file = "extracted_data.json"
excel_file = "orders.xlsx"

# Load JSON
if not os.path.exists(json_file):
    print("No extracted_data.json found. Run extract.py first.")
    exit()

with open(json_file, "r", encoding="utf-8") as f:
    data = json.load(f)

# Correct headers
HEADERS = [
    "customer_name", "phone_number", "email", "service_address",
    "job_type", "problem_description", "urgency", "preferred_time",
    "technician_preference", "price_reference", "notes"
]

# Create Excel if missing
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(HEADERS)
    wb.save(excel_file)

# Load workbook
wb = load_workbook(excel_file)
ws = wb.active

# Fix headers if needed
existing_headers = [cell.value for cell in ws[1]]
if existing_headers != HEADERS:
    print("⚙Fixing headers...")
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    wb.save(excel_file)

# Append new row
ws.append([data.get(key) for key in HEADERS])

# 🔧 Auto-fit column widths
for col_num, header in enumerate(HEADERS, start=1):
    max_length = len(header)
    for cell in ws[get_column_letter(col_num)]:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

wb.save(excel_file)

print(f"Data appended and formatted in {excel_file}")

