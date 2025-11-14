# ai_notetaker.py
# Full pipeline: transcribe audio → extract structured data → save to Excel

import sys
import os
import json
import sqlite3
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

print(">>> Running ai_notetaker with:", sys.executable)
print(">>> sys.path contains openpyxl?", any("openpyxl" in p for p in sys.path))

# ─── CONFIG ──────────────────────────────────────────────
AUDIO_FILE = "sample_call.m4a"       # change if your file has another name
EXCEL_FILE = "orders.xlsx"
MODEL_TRANSCRIBE = "whisper-1"
MODEL_EXTRACT = "gpt-4o-mini"


HEADERS = [
    "customer_name", "phone_number", "email", "service_address",
    "job_type", "problem_description", "urgency", "preferred_time",
    "technician_preference", "price_reference", "notes",
    "audio_file", "created_at", "transcript_preview"
]

# ─────────────────────────────────────────────────────────

load_dotenv()
client = OpenAI()

# 1 Transcription
if not os.path.exists(AUDIO_FILE):
    print(f"Audio file {AUDIO_FILE} not found.")
    exit()

print("Transcribing call...")
with open(AUDIO_FILE, "rb") as audio:
    transcript = client.audio.transcriptions.create(
        model=MODEL_TRANSCRIBE,
        file=audio
    )

# Full transcript text
text = transcript.text
print("Transcription complete.")

# ─── ADD METADATA ─────────────────────────────────────────────
from datetime import datetime

created_at = datetime.now().isoformat(timespec="seconds")
transcript_preview = text[:200]  # first 200 characters of transcript

metadata = {
    "audio_file": AUDIO_FILE,
    "created_at": created_at,
    "transcript_preview": transcript_preview
}


# 2️ Extraction
print("Extracting structured data…")
response = client.chat.completions.create(
    model=MODEL_EXTRACT,
    response_format={"type": "json_object"},
    temperature=0,
    messages=[
        {"role": "system", "content": "Extract structured job data in JSON."},
        {"role": "user", "content": f"""
From the following customer call transcript, extract the fields below.

Return ONLY valid JSON.

Fields:
{', '.join(HEADERS)}

Transcript:
{text}
"""}
    ]
)

data = json.loads(response.choices[0].message.content)
print("Data extracted.")

# 3️ Save to Excel
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# fix headers if needed
existing_headers = [c.value for c in ws[1]]
if existing_headers != HEADERS:
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

# append row
row = []
for h in HEADERS:
    if h in data:
        row.append(data.get(h))
    else:
        row.append(metadata.get(h))

ws.append(row)


# auto-fit columns
for col_num, header in enumerate(HEADERS, start=1):
    max_len = len(header)
    for cell in ws[get_column_letter(col_num)]:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_num)].width = max_len + 2

# Save to Excel
wb.save(EXCEL_FILE)
print(f"Saved structured data to {EXCEL_FILE}")

# ─── SAVE TO DATABASE ─────────────────────────────────────────────

conn = sqlite3.connect("jobs.db")
cursor = conn.cursor()

cursor.execute("""
INSERT INTO jobs (
    customer_name, phone_number, email, service_address,
    job_type, problem_description, urgency, preferred_time,
    technician_preference, price_reference, notes,
    audio_file, created_at, transcript_preview
)
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
""", [
    data.get("customer_name"),
    data.get("phone_number"),
    data.get("email"),
    data.get("service_address"),
    data.get("job_type"),
    data.get("problem_description"),
    data.get("urgency"),
    data.get("preferred_time"),
    data.get("technician_preference"),
    data.get("price_reference"),
    data.get("notes"),
    metadata["audio_file"],
    metadata["created_at"],
    metadata["transcript_preview"]
])

conn.commit()
conn.close()

print("Saved job to SQLite database (jobs.db)")

# FINAL MESSAGE
print("Workflow complete!")

def process_audio_file(path):
    global AUDIO_FILE
    AUDIO_FILE = path
    process_audio(AUDIO_FILE)   # run your existing workflow


if __name__ == "__main__":
    process_audio_file(AUDIO_FILE)
